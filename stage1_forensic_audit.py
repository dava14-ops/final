#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
stage1_forensic_audit.py

Forensic audit официального файла Росстата `cena_sx_07-2026.xlsx`.

Цель: ответить на 4 критических вопроса ДО написания production parser:
1. Почему листы 1.1/1.2 остановились на 2021?
2. Почему лист 6.2 дал 0 строк?
3. Где в действительности находятся ФО/субъекты?
4. Как реально устроены строки территория → культура → месяцы?

Принципы:
- Никаких трансформаций данных
- Прямой доступ к openpyxl (merged cells, координаты)
- Полный вывод структуры для каждого листа
- Детальный разбор листов 6.1 и 6.2 (первые 100 строк)

Использование:
    python stage1_forensic_audit.py --input cena_sx_07-2026.xlsx
    python stage1_forensic_audit.py --input cena_sx_07-2026.xlsx --verbose
"""

import argparse
import logging
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.cell.cell import Cell
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl не установлен. Установите: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Логирование
# ---------------------------------------------------------------------------

def setup_logging(verbose: bool = False) -> logging.Logger:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(message)s",
    )
    return logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Утилиты
# ---------------------------------------------------------------------------

def safe_str(value: Any) -> str:
    """Безопасное преобразование значения в строку."""
    if value is None:
        return ""
    return str(value).strip()


def is_year_like(value: Any) -> bool:
    """Проверить, похоже ли значение на год (2010-2030)."""
    s = safe_str(value)
    if re.match(r"^20\d{2}$", s):
        return True
    # Проверяем числовое значение
    try:
        v = float(s)
        if 2000 <= v <= 2030:
            return True
    except (ValueError, TypeError):
        pass
    return False


def is_month_like(value: Any) -> bool:
    """Проверить, похоже ли значение на месяц."""
    s = safe_str(value).lower()
    months = {
        "январь", "февраль", "март", "апрель", "май", "июнь",
        "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
        "янв", "фев", "мар", "апр", "май", "июн", "июл", "авг",
        "сен", "окт", "ноя", "дек",
    }
    return s in months


def extract_okpd2(value: Any) -> Optional[str]:
    """Извлечь ОКПД2 код из значения."""
    s = safe_str(value)
    # Паттерн ОКПД2: две цифры, точка, две цифры, далее опционально
    pattern = re.compile(
        r"(\d{2}\.\d{2}(?:\.\d{1,3})?(?:\.\d{3})?(?:\.АГ)?)"
    )
    match = pattern.search(s)
    if match:
        return match.group(1)
    return None


def classify_territory(value: Any) -> str:
    """Классифицировать территорию."""
    s = safe_str(value).lower()

    if "российская федерация" in s:
        return "national"

    # Федеральные округа
    fo_pattern = re.compile(
        r"(центральный|северо-западный|южный|северо-кавказский|"
        r"приволжский|уральский|сибирский|дальневосточный)\s+"
        r"федеральный округ",
        re.IGNORECASE,
    )
    if fo_pattern.search(s):
        return "federal_district"

    # Субъекты РФ (области, края, республики)
    subject_patterns = [
        r"область",
        r"край",
        r"республика",
        r"автономный округ",
        r"город федерального значения",
        r"москва",
        r"санкт-петербург",
        r"севастополь",
    ]
    for pattern in subject_patterns:
        if re.search(pattern, s):
            return "subject"

    return "unknown"


def print_separator(char: str = "=", length: int = 80) -> None:
    """Вывести разделитель."""
    print(char * length)


def print_header(title: str) -> None:
    """Вывести заголовок."""
    print()
    print_separator()
    print(f"  {title}")
    print_separator()


# ---------------------------------------------------------------------------
# Аудит одного листа
# ---------------------------------------------------------------------------

def audit_sheet(
    ws,
    sheet_name: str,
    max_sample_rows: int = 100,
    verbose: bool = False,
) -> Dict[str, Any]:
    """
    Полный аудит одного листа.

    Возвращает словарь с результатами аудита.
    """
    logger = logging.getLogger(__name__)

    result = {
        "sheet_name": sheet_name,
        "dimensions": None,
        "max_row": 0,
        "max_col": 0,
        "merged_cells": [],
        "header_candidates": [],
        "year_columns": [],
        "month_columns": [],
        "territory_counts": {
            "national": 0,
            "federal_district": 0,
            "subject": 0,
            "unknown": 0,
        },
        "okpd2_counts": {
            "with_okpd2": 0,
            "without_okpd2": 0,
        },
        "detected_years": [],
        "detected_months": [],
        "data_sample": [],
        "issues": [],
    }

    # 1. Размеры
    result["dimensions"] = ws.dimensions
    result["max_row"] = ws.max_row
    result["max_col"] = ws.max_column

    print_header(f"ЛИСТ: {sheet_name}")
    print(f"  Размеры: {ws.dimensions}")
    print(f"  Строк: {ws.max_row}, Колонок: {ws.max_column}")

    # 2. Merged cells
    merged = list(ws.merged_cells.ranges)
    result["merged_cells"] = [str(m) for m in merged]

    if merged:
        print(f"\n  MERGED CELLS: {len(merged)}")
        for i, m in enumerate(merged[:20]):
            print(f"    {m}")
        if len(merged) > 20:
            print(f"    ... и ещё {len(merged) - 20}")
    else:
        print("\n  MERGED CELLS: нет")

    # 3. Поиск заголовков (первые 20 строк)
    print(f"\n  HEADER CANDIDATES (первые 20 строк):")
    print_separator("-", 80)

    header_row_idx = None
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = safe_str(cell.value)
            if val:
                row_values.append(val)

        if not row_values:
            continue

        # Проверяем, содержит ли строка годы или месяцы
        years_in_row = [v for v in row_values if is_year_like(v)]
        months_in_row = [v for v in row_values if is_month_like(v)]

        is_header = len(years_in_row) >= 2 or len(months_in_row) >= 2

        marker = " ← HEADER" if is_header else ""
        print(f"  Row {row_idx:3d}: {row_values[:8]}{'...' if len(row_values) > 8 else ''}{marker}")

        result["header_candidates"].append({
            "row": row_idx,
            "values": row_values[:15],
            "years_found": years_in_row,
            "months_found": months_in_row,
            "is_header": is_header,
        })

        if is_header and header_row_idx is None:
            header_row_idx = row_idx

    # 4. Определение годовых и месячных колонок
    print(f"\n  TIME STRUCTURE:")
    print_separator("-", 80)

    if header_row_idx is not None:
        print(f"  Заголовок найден в строке {header_row_idx}")

        # Сканируем строку заголовка
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            val = safe_str(cell.value)
            col_letter = get_column_letter(col_idx)

            if is_year_like(val):
                year = int(float(val))
                result["year_columns"].append({
                    "col_idx": col_idx,
                    "col_letter": col_letter,
                    "year": year,
                })
                if year not in result["detected_years"]:
                    result["detected_years"].append(year)

            elif is_month_like(val):
                result["month_columns"].append({
                    "col_idx": col_idx,
                    "col_letter": col_letter,
                    "month_name": val,
                })
                if val not in result["detected_months"]:
                    result["detected_months"].append(val)

        if result["year_columns"]:
            years = [yc["year"] for yc in result["year_columns"]]
            print(f"  Годовые колонки: {len(years)}")
            print(f"  Диапазон: {min(years)} – {max(years)}")
            print(f"  Годы: {sorted(years)}")
        else:
            print(f"  Годовые колонки: НЕ НАЙДЕНЫ")

        if result["month_columns"]:
            months = [mc["month_name"] for mc in result["month_columns"]]
            print(f"  Месячные колонки: {len(months)}")
            print(f"  Месяцы: {months}")
        else:
            print(f"  Месячные колонки: НЕ НАЙДЕНЫ")
    else:
        print(f"  Заголовок НЕ НАЙДЕН в первых 20 строках")
        result["issues"].append("Header not found in first 20 rows")

    # 5. Сэмпл данных (первые max_sample_rows строк после заголовка)
    print(f"\n  DATA SAMPLE (первые {max_sample_rows} строк):")
    print_separator("-", 80)

    start_row = header_row_idx + 1 if header_row_idx else 1
    end_row = min(start_row + max_sample_rows, ws.max_row + 1)

    for row_idx in range(start_row, end_row):
        row_data = {}
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = safe_str(cell.value)
            if val:
                col_letter = get_column_letter(col_idx)
                row_data[col_letter] = val

        if row_data:
            result["data_sample"].append({
                "excel_row": row_idx,
                "cells": row_data,
            })

            # Классифицируем территорию
            first_cell_val = safe_str(ws.cell(row=row_idx, column=1).value)
            territory_type = classify_territory(first_cell_val)
            result["territory_counts"][territory_type] += 1

            # Проверяем ОКПД2
            has_okpd2 = False
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                cell_val = safe_str(ws.cell(row=row_idx, column=col_idx).value)
                if extract_okpd2(cell_val):
                    has_okpd2 = True
                    break

            if has_okpd2:
                result["okpd2_counts"]["with_okpd2"] += 1
            else:
                result["okpd2_counts"]["without_okpd2"] += 1

    # Выводим сэмпл
    for sample in result["data_sample"][:30]:
        row_str = ", ".join(
            f"{k}={v[:30]}{'...' if len(v) > 30 else ''}"
            for k, v in list(sample["cells"].items())[:6]
        )
        print(f"  Row {sample['excel_row']:4d}: {row_str}")

    if len(result["data_sample"]) > 30:
        print(f"  ... и ещё {len(result['data_sample']) - 30} строк")

    # 6. Структура территорий
    print(f"\n  TERRITORY STRUCTURE:")
    print_separator("-", 80)
    for geo_type, count in result["territory_counts"].items():
        print(f"  {geo_type}: {count}")

    # 7. Структура ОКПД2
    print(f"\n  OKPD2 STRUCTURE:")
    print_separator("-", 80)
    print(f"  С ОКПД2: {result['okpd2_counts']['with_okpd2']}")
    print(f"  Без ОКПД2: {result['okpd2_counts']['without_okpd2']}")

    # 8. Issues
    if result["issues"]:
        print(f"\n  ⚠️ ISSUES:")
        for issue in result["issues"]:
            print(f"    - {issue}")

    return result


# ---------------------------------------------------------------------------
# Специальный аудит листов 6.1 и 6.2
# ---------------------------------------------------------------------------

def audit_sheet_6x(
    ws,
    sheet_name: str,
    verbose: bool = False,
) -> Dict[str, Any]:
    """
    Углублённый аудит листов 6.1 и 6.2.

    Эти листы содержат РФ + ФО + субъекты, и нам нужно понять
    реальную структуру строк.
    """
    logger = logging.getLogger(__name__)

    print_header(f"УГЛУБЛЁННЫЙ АУДИТ ЛИСТА: {sheet_name}")

    result = {
        "sheet_name": sheet_name,
        "total_rows": ws.max_row,
        "total_cols": ws.max_column,
        "merged_cells": [],
        "first_100_rows": [],
        "territory_hierarchy": [],
        "issues": [],
    }

    # Merged cells
    merged = list(ws.merged_cells.ranges)
    result["merged_cells"] = [str(m) for m in merged]
    print(f"  Merged cells: {len(merged)}")

    # Первые 100 строк — полный разбор
    print(f"\n  ПЕРВЫЕ 100 СТРОК (полный разбор):")
    print_separator("-", 80)

    current_territory = ""
    territory_hierarchy = []

    for row_idx in range(1, min(101, ws.max_row + 1)):
        row_data = {}
        for col_idx in range(1, min(ws.max_column + 1, 25)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = safe_str(cell.value)
            if val:
                col_letter = get_column_letter(col_idx)
                row_data[col_letter] = val

        if not row_data:
            continue

        result["first_100_rows"].append({
            "excel_row": row_idx,
            "cells": row_data,
        })

        # Определяем территорию
        first_val = safe_str(ws.cell(row=row_idx, column=1).value)
        territory_type = classify_territory(first_val)

        if territory_type in ("national", "federal_district", "subject"):
            current_territory = first_val
            territory_hierarchy.append({
                "row": row_idx,
                "territory": first_val,
                "type": territory_type,
            })

        # Выводим строку
        row_str = ", ".join(
            f"{k}={v[:40]}{'...' if len(v) > 40 else ''}"
            for k, v in list(row_data.items())[:8]
        )
        print(f"  Row {row_idx:4d}: {row_str}")

    result["territory_hierarchy"] = territory_hierarchy

    # Анализируем иерархию
    print(f"\n  TERRITORY HIERARCHY (обнаружено {len(territory_hierarchy)} территорий):")
    print_separator("-", 80)

    for item in territory_hierarchy[:50]:
        print(f"  Row {item['row']:4d}: [{item['type']}] {item['territory'][:60]}")

    if len(territory_hierarchy) > 50:
        print(f"  ... и ещё {len(territory_hierarchy) - 50}")

    # Подсчёт типов
    type_counts = {}
    for item in territory_hierarchy:
        t = item["type"]
        type_counts[t] = type_counts.get(t, 0) + 1

    print(f"\n  Подсчёт типов территорий:")
    for t, c in sorted(type_counts.items()):
        print(f"    {t}: {c}")

    return result


# ---------------------------------------------------------------------------
# Основной аудит
# ---------------------------------------------------------------------------

def audit_workbook(input_path: Path, verbose: bool = False) -> Dict[str, Any]:
    """
    Полный аудит workbook.
    """
    logger = logging.getLogger(__name__)

    print_separator("=")
    print("  FORENSIC WORKBOOK AUDIT")
    print(f"  Файл: {input_path}")
    print(f"  Размер: {input_path.stat().st_size:,} байт")
    print_separator("=")

    # Открываем workbook
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
    logger.info(f"Листы в файле: {wb.sheetnames}")

    results = {
        "file": str(input_path),
        "sheets": {},
        "issues": [],
    }

    for sheet_name in wb.sheetnames:
        # Пропускаем служебные листы
        if sheet_name.lower() in ("содержание", "оглавление", "contents"):
            print(f"\nПропуск служебного листа: '{sheet_name}'")
            continue

        ws = wb[sheet_name]

        # Для листов 6.1 и 6.2 — углублённый аудит
        if sheet_name in ("6.1", "6.2", "7.1"):
            result = audit_sheet_6x(ws, sheet_name, verbose)
        else:
            result = audit_sheet(ws, sheet_name, verbose=verbose)

        results["sheets"][sheet_name] = result

    # Итоговый отчёт
    print_header("ИТОГОВЫЙ ОТЧЁТ")

    print("\n  Сводка по листам:")
    print_separator("-", 80)
    print(f"  {'Лист':<10} {'Строк':>8} {'Колонок':>8} {'Записей':>10} {'Проблемы'}")
    print_separator("-", 80)

    for sheet_name, result in results["sheets"].items():
        n_rows = result.get("max_row", result.get("total_rows", 0))
        n_cols = result.get("max_col", result.get("total_cols", 0))
        n_records = len(result.get("data_sample", result.get("first_100_rows", [])))
        issues = result.get("issues", [])
        issues_str = ", ".join(issues) if issues else "OK"
        print(f"  {sheet_name:<10} {n_rows:>8} {n_cols:>8} {n_records:>10} {issues_str}")

    # Критические вопросы
    print_header("КРИТИЧЕСКИЕ ВОПРОСЫ ДЛЯ РАЗРЕШЕНИЯ")

    print("\n  1. Листы 1.1 и 1.2:")
    for sn in ("1.1", "1.2"):
        if sn in results["sheets"]:
            sheet = results["sheets"][sn]
            years = sheet.get("detected_years", [])
            if years:
                print(f"     {sn}: обнаружены годы {min(years)}–{max(years)}")
                if max(years) < 2025:
                    print(f"     ⚠️ ГОДЫ 2022–2025 НЕ ОБНАРУЖЕНЫ!")
                    print(f"     Возможные причины:")
                    print(f"       - Годы в merged cells")
                    print(f"       - Годы в другой строке заголовка")
                    print(f"       - Файл действительно содержит только до {max(years)}")
            else:
                print(f"     {sn}: ГОДЫ НЕ ОБНАРУЖЕНЫ")

    print("\n  2. Лист 6.2:")
    if "6.2" in results["sheets"]:
        sheet = results["sheets"]["6.2"]
        n_records = len(sheet.get("first_100_rows", []))
        if n_records == 0:
            print(f"     ⚠️ ЛИСТ 6.2 ПУСТ ИЛИ НЕ ЧИТАЕТСЯ!")
        else:
            print(f"     Обнаружено {n_records} строк в первых 100")

    print("\n  3. Федеральные округа:")
    fo_found = False
    for sn, sheet in results["sheets"].items():
        counts = sheet.get("territory_counts", {})
        if counts.get("federal_district", 0) > 0:
            fo_found = True
            print(f"     Лист {sn}: найдено {counts['federal_district']} ФО")

        # Проверяем иерархию
        hierarchy = sheet.get("territory_hierarchy", [])
        for item in hierarchy:
            if item["type"] == "federal_district":
                fo_found = True
                print(f"     Лист {sn}, строка {item['row']}: {item['territory'][:50]}")

    if not fo_found:
        print(f"     ⚠️ ФЕДЕРАЛЬНЫЕ ОКРУГА НЕ ОБНАРУЖЕНЫ!")
        print(f"     Возможные причины:")
        print(f"       - ФО не представлены в файле")
        print(f"       - ФО в другой колонке")
        print(f"       - Паттерн распознавания ФО некорректен")

    return results


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Forensic audit файла Росстата cena_sx_07-2026.xlsx"
    )
    parser.add_argument(
        "--input",
        type=str,
        default="cena_sx_07-2026.xlsx",
        help="Путь к входному XLSX файлу",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Подробный вывод",
    )

    args = parser.parse_args()
    logger = setup_logging(args.verbose)

    input_path = Path(args.input)
    if not input_path.exists():
        logger.error(f"Файл не найден: {input_path}")
        return 1

    try:
        results = audit_workbook(input_path, verbose=args.verbose)
    except Exception as exc:
        logger.error(f"Ошибка аудита: {exc}")
        import traceback
        traceback.print_exc()
        return 1

    print_header("АУДИТ ЗАВЕРШЁН")
    return 0


if __name__ == "__main__":
    sys.exit(main())