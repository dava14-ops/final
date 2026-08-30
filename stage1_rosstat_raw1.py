#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
stage1_rosstat_raw.py (v2.0 - Production Grade)

Детерминированный парсер официального файла Росстата `cena_sx_07-2026.xlsx`.
Исправлены все баги, выявленные в ходе форензик-аудита:
1. Сноски у годов ('2022 1') в листах 1.1/1.2.
2. Сдвиг колонок в листе 6.2.
3. Блочная структура (State-Machine) в листах 6.1/7.1.
4. Строгое разделение статусов пропусков ('...' vs '-').

Использование:
    python stage1_rosstat_raw.py --input cena_sx_07-2026.xlsx
"""

import argparse
import logging
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    import pandas as pd
except ImportError:
    print("ERROR: Требуется установить openpyxl и pandas: pip install openpyxl pandas")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Конфигурация и паттерны
# ---------------------------------------------------------------------------

MONTH_MAP = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}

# Точный regex для ОКПД2 (например: 01.11.1, 01.11.95, 01.11.12.001.АГ)
OKPD2_REGEX = re.compile(r"^\d{2}\.\d{2}(?:\.\d{1,3})?(?:\.\d{3})?(?:\.АГ)?$")

# Metadata для листов
SHEET_CONFIG = {
    "1.1": {
        "year": None,
        "freq": "annual",
        "timing": "annual_average",
        "geo": "national",
        "parser": "annual_1_1",
    },
    "1.2": {
        "year": None,
        "freq": "annual",
        "timing": "end_of_year",
        "geo": "national",
        "parser": "annual_1_2",
    },
    "2": {
        "year": 2021,
        "freq": "monthly",
        "timing": "end_of_period",
        "geo": "national",
        "parser": "national_monthly",
    },
    "3": {
        "year": 2022,
        "freq": "monthly",
        "timing": "end_of_period",
        "geo": "national",
        "parser": "national_monthly",
    },
    "4": {
        "year": 2023,
        "freq": "monthly",
        "timing": "end_of_period",
        "geo": "national",
        "parser": "national_monthly",
    },
    "5": {
        "year": 2024,
        "freq": "monthly",
        "timing": "end_of_period",
        "geo": "national",
        "parser": "national_monthly",
    },
    "6.1": {
        "year": 2025,
        "freq": "monthly",
        "timing": "end_of_period",
        "geo": "mixed",
        "parser": "mixed_monthly",
    },
    "6.2": {
        "year": 2025,
        "freq": "annual",
        "timing": "annual_average",
        "geo": "mixed",
        "parser": "mixed_6_2",
    },
    "7.1": {
        "year": 2026,
        "freq": "monthly",
        "timing": "end_of_period",
        "geo": "mixed",
        "parser": "mixed_monthly",
    },
}

# ---------------------------------------------------------------------------
# Утилиты
# ---------------------------------------------------------------------------


def setup_logging():
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s | %(levelname)-7s | %(message)s", datefmt="%H:%M:%S"
    )


def extract_year_from_header(val) -> Optional[int]:
    """Извлекает год из заголовка, игнорируя сноски (например, '2022 1' -> 2022)."""
    if val is None:
        return None
    s = str(val).strip()
    match = re.match(r"^(20\d{2})", s)
    if match:
        return int(match.group(1))
    try:
        v = int(float(s))
        if 2000 <= v <= 2030:
            return v
    except (ValueError, TypeError):
        pass
    return None


def clean_and_classify_value(val) -> Tuple[Optional[float], str]:
    """Классифицирует значение ячейки, сохраняя семантику пропусков."""
    if val is None:
        return None, "structurally_missing"
    s = str(val).strip()
    if not s:
        return None, "structurally_missing"
    if s == "...":
        return None, "confidential"
    if s == "-":
        return None, "unavailable"

    cleaned = s.replace(" ", "").replace("\u00a0", "").replace(",", ".")
    try:
        return float(cleaned), "observed"
    except ValueError:
        return None, "structurally_missing"


def classify_territory(name: str) -> str:
    """Определяет уровень территории."""
    if not name:
        return "unknown"
    s = str(name).strip().lower()
    if "российская федерация" in s:
        return "national"
    if "федеральный округ" in s or s.endswith("фо"):
        return "federal_district"

    subject_kw = [
        "область",
        "край",
        "республика",
        "автономный округ",
        "москва",
        "санкт-петербург",
        "севастополь",
    ]
    if any(kw in s for kw in subject_kw):
        return "subject"
    if "г.севастополь" in s or "город федерального значения" in s:
        return "subject"

    return "unknown"


@dataclass
class RawPriceRecord:
    source_file: str
    source_sheet: str
    source_row_excel: int
    source_column_excel: str
    year: Optional[int]
    month: Optional[int]
    frequency: str
    geography_level: str
    region_name: str
    region_source_code: Optional[str]
    product_name_raw: str
    okpd2_raw: Optional[str]
    raw_value: str
    value_numeric: Optional[float]
    value_status: str
    price_timing: str
    unit_raw: str


# ---------------------------------------------------------------------------
# Парсеры (State-Machines)
# ---------------------------------------------------------------------------


def parse_annual_1_1(ws, source_file) -> List[RawPriceRecord]:
    """Лист 1.1: Годовые национальные. Колонка A = продукт, B-Q = годы."""
    records = []
    year_cols = {c: extract_year_from_header(ws.cell(4, c).value) for c in range(2, 18)}
    year_cols = {c: y for c, y in year_cols.items() if y is not None}

    for r in range(5, ws.max_row + 1):
        prod = str(ws.cell(r, 1).value or "").strip()
        if not prod or prod.lower().startswith(("источник", "примечание", "1 ")):
            continue

        for c, y in year_cols.items():
            val, status = clean_and_classify_value(ws.cell(r, c).value)
            records.append(
                RawPriceRecord(
                    source_file,
                    "1.1",
                    r,
                    get_column_letter(c),
                    y,
                    None,
                    "annual",
                    "national",
                    "Российская Федерация",
                    None,
                    prod,
                    None,
                    str(ws.cell(r, c).value or "").strip(),
                    val,
                    status,
                    "annual_average",
                    "RUB/ton",
                )
            )
    return records


def parse_annual_1_2(ws, source_file) -> List[RawPriceRecord]:
    """Лист 1.2: Годовые национальные. A = продукт, B = ОКПД2, C-R = годы."""
    records = []
    year_cols = {c: extract_year_from_header(ws.cell(4, c).value) for c in range(3, 19)}
    year_cols = {c: y for c, y in year_cols.items() if y is not None}

    for r in range(5, ws.max_row + 1):
        prod = str(ws.cell(r, 1).value or "").strip()
        okpd = str(ws.cell(r, 2).value or "").strip()
        if not prod or prod.lower().startswith(("источник", "примечание", "1 ")):
            continue
        okpd = okpd if OKPD2_REGEX.match(okpd) else None

        for c, y in year_cols.items():
            val, status = clean_and_classify_value(ws.cell(r, c).value)
            records.append(
                RawPriceRecord(
                    source_file,
                    "1.2",
                    r,
                    get_column_letter(c),
                    y,
                    None,
                    "annual",
                    "national",
                    "Российская Федерация",
                    None,
                    prod,
                    okpd,
                    str(ws.cell(r, c).value or "").strip(),
                    val,
                    status,
                    "end_of_year",
                    "RUB/ton",
                )
            )
    return records


def parse_national_monthly(ws, sheet_name, year, source_file) -> List[RawPriceRecord]:
    """Листы 2-5: Месячные национальные. A = продукт, B = ОКПД2, C-N = месяцы."""
    records = []
    header_row = next(
        (
            r
            for r in range(1, 10)
            if "январь" in str(ws.cell(r, 2).value or "").lower()
            or "январь" in str(ws.cell(r, 3).value or "").lower()
        ),
        None,
    )
    if not header_row:
        raise ValueError(f"Sheet {sheet_name}: Header not found")

    month_cols = {}
    for c in range(1, 15):
        m_name = str(ws.cell(header_row, c).value or "").strip().lower()
        if m_name in MONTH_MAP:
            month_cols[c] = MONTH_MAP[m_name]

    for r in range(header_row + 1, ws.max_row + 1):
        prod = str(ws.cell(r, 1).value or "").strip()
        okpd = str(ws.cell(r, 2).value or "").strip()
        if not prod or prod.lower().startswith(("источник", "примечание", "1 ")):
            continue
        okpd = okpd if OKPD2_REGEX.match(okpd) else None

        for c, m in month_cols.items():
            val, status = clean_and_classify_value(ws.cell(r, c).value)
            records.append(
                RawPriceRecord(
                    source_file,
                    sheet_name,
                    r,
                    get_column_letter(c),
                    year,
                    m,
                    "monthly",
                    "national",
                    "Российская Федерация",
                    None,
                    prod,
                    okpd,
                    str(ws.cell(r, c).value or "").strip(),
                    val,
                    status,
                    "end_of_period",
                    "RUB/ton",
                )
            )
    return records


def parse_mixed_monthly(ws, sheet_name, year, source_file) -> List[RawPriceRecord]:
    """Листы 6.1, 7.1: Блочная структура (State-Machine)."""
    records = []
    header_row = next(
        (r for r in range(1, 10) if "январь" in str(ws.cell(r, 3).value or "").lower()), None
    )
    if not header_row:
        raise ValueError(f"Sheet {sheet_name}: Header not found")

    month_cols = {}
    for c in range(1, 15):
        m_name = str(ws.cell(header_row, c).value or "").strip().lower()
        if m_name in MONTH_MAP:
            month_cols[c] = MONTH_MAP[m_name]

    current_crop, current_okpd2 = None, None

    for r in range(header_row + 1, ws.max_row + 1):
        cell_a = str(ws.cell(r, 1).value or "").strip()
        cell_b = str(ws.cell(r, 2).value or "").strip()

        if not cell_a and not cell_b:
            continue
        if cell_a.lower().startswith(("источник", "примечание", "1 ")):
            continue

        # 1. Это заголовок блока культуры? (A=Name, B=OKPD2)
        if cell_b and OKPD2_REGEX.match(cell_b):
            current_crop = cell_a
            current_okpd2 = cell_b
            continue

        # 2. Это строка с данными территории?
        if current_crop and cell_a:
            geo = classify_territory(cell_a)
            for c, m in month_cols.items():
                val, status = clean_and_classify_value(ws.cell(r, c).value)
                records.append(
                    RawPriceRecord(
                        source_file,
                        sheet_name,
                        r,
                        get_column_letter(c),
                        year,
                        m,
                        "monthly",
                        geo,
                        cell_a,
                        None,
                        current_crop,
                        current_okpd2,
                        str(ws.cell(r, c).value or "").strip(),
                        val,
                        status,
                        "end_of_period",
                        "RUB/ton",
                    )
                )
    return records


def parse_mixed_6_2(ws, source_file) -> List[RawPriceRecord]:
    """Лист 6.2: Сдвинутая блочная структура (B=Name, C=OKPD2, D=Value)."""
    records = []
    header_row = next(
        (r for r in range(1, 10) if "окпд2" in str(ws.cell(r, 3).value or "").lower()), None
    )
    if not header_row:
        raise ValueError("Sheet 6.2: Header not found")

    current_crop, current_okpd2 = None, None

    for r in range(header_row + 1, ws.max_row + 1):
        cell_b = str(ws.cell(r, 2).value or "").strip()
        cell_c = str(ws.cell(r, 3).value or "").strip()

        if not cell_b and not cell_c:
            continue
        if cell_b.lower().startswith(("источник", "примечание", "1 ")):
            continue

        # Заголовок блока (B=Name, C=OKPD2)
        if cell_c and OKPD2_REGEX.match(cell_c):
            current_crop = cell_b
            current_okpd2 = cell_c
            continue

        # Строка территории (B=Territory, D=Value)
        if current_crop and cell_b:
            geo = classify_territory(cell_b)
            val, status = clean_and_classify_value(ws.cell(r, 4).value)
            records.append(
                RawPriceRecord(
                    source_file,
                    "6.2",
                    r,
                    "D",
                    2025,
                    None,
                    "annual",
                    geo,
                    cell_b,
                    None,
                    current_crop,
                    current_okpd2,
                    str(ws.cell(r, 4).value or "").strip(),
                    val,
                    status,
                    "annual_average",
                    "RUB/ton",
                )
            )
    return records


# ---------------------------------------------------------------------------
# Main Pipeline
# ---------------------------------------------------------------------------


def main():
    setup_logging()
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="cena_sx_07-2026.xlsx")
    parser.add_argument("--output", default="rosstat_prices_raw_long_v2.csv")
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        logging.error(f"Файл не найден: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    all_records = []
    sheet_stats = {}

    parsers = {
        "annual_1_1": lambda ws, cfg: parse_annual_1_1(ws, path.name),
        "annual_1_2": lambda ws, cfg: parse_annual_1_2(ws, path.name),
        "national_monthly": lambda ws, cfg: parse_national_monthly(
            ws, cfg["sheet"], cfg["year"], path.name
        ),
        "mixed_monthly": lambda ws, cfg: parse_mixed_monthly(
            ws, cfg["sheet"], cfg["year"], path.name
        ),
        "mixed_6_2": lambda ws, cfg: parse_mixed_6_2(ws, path.name),
    }

    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            logging.warning(f"Лист {sheet_name} отсутствует в файле!")
            continue

        logging.info(f"Парсинг листа {sheet_name}...")
        ws = wb[sheet_name]
        cfg["sheet"] = sheet_name

        try:
            records = parsers[cfg["parser"]](ws, cfg)
        except Exception as e:
            logging.error(f"КРИТИЧЕСКАЯ ОШИБКА при парсинге {sheet_name}: {e}")
            sys.exit(1)

        # 🔴 STRICT ASSERTION: Ожидаемый лист не может быть пустым
        if len(records) == 0:
            logging.error(
                f"❌ ASSERTION FAILED: Лист {sheet_name} дал 0 записей! Pipeline остановлен."
            )
            sys.exit(1)

        all_records.extend(records)
        sheet_stats[sheet_name] = len(records)
        logging.info(f"  ✅ {sheet_name}: {len(records)} записей")

    # Сохранение
    df = pd.DataFrame([asdict(r) for r in all_records])
    df.to_csv(args.output, index=False, encoding="utf-8-sig")

    logging.info("=" * 60)
    logging.info(f"ИТОГО: {len(df):,} записей сохранено в {args.output}")
    logging.info("Статистика по листам:")
    for s, c in sheet_stats.items():
        logging.info(f"  {s}: {c}")
    logging.info("=" * 60)


if __name__ == "__main__":
    main()
